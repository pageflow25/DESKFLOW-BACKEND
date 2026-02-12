"""
Script para gerar relatório de orçamento PageFlow em Excel
"""
import json
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def estilizar_cabecalho(ws, row=1):
    """Aplica estilo ao cabeçalho"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def ajustar_largura_colunas(ws):
    """Ajusta automaticamente a largura das colunas"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def gerar_relatorio_excel(arquivo_json: str, arquivo_saida: str = None):
    """
    Gera um relatório em Excel a partir do arquivo JSON PageFlow
    
    Args:
        arquivo_json: Caminho do arquivo JSON
        arquivo_saida: Caminho do arquivo Excel de saída (opcional)
    """
    
    # Ler o arquivo JSON
    with open(arquivo_json, 'r', encoding='utf-8') as f:
        dados = json.load(f)
    
    data = dados.get('data', {})
    itens = data.get('itens', [])
    itens_validos = [item for item in itens if item and item.get('id_produto')]
    
    # Criar workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove a planilha padrão
    
    # ===== ABA 1: RESUMO GERAL =====
    ws_resumo = wb.create_sheet("Resumo Geral")
    
    resumo_data = [
        ["RELATÓRIO DE ORÇAMENTO PAGEFLOW", ""],
        ["Data de Geração", datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
        ["", ""],
        ["INFORMAÇÕES GERAIS", ""],
        ["Identificador", dados.get('identifier', 'N/A')],
        ["ID Escola", data.get('id_escola', 'N/A')],
        ["ID Cliente", data.get('id_cliente', 'N/A')],
        ["ID Vendedor", data.get('id_vendedor', 'N/A')],
        ["ID Forma de Pagamento", data.get('id_forma_pagamento', 'N/A')],
        ["", ""],
        ["ESTATÍSTICAS", ""],
        ["Total de Itens", len(itens)],
        ["Itens Válidos", len(itens_validos)],
        ["Itens Vazios", len(itens) - len(itens_validos)],
        ["Quantidade Total de Produtos", sum(item.get('quantidade', 0) for item in itens_validos)],
        ["Produtos Únicos", len(set(item.get('id_produto') for item in itens_validos))],
    ]
    
    for row_data in resumo_data:
        ws_resumo.append(row_data)
    
    # Estilizar títulos
    ws_resumo['A1'].font = Font(bold=True, size=14, color="366092")
    ws_resumo['A4'].font = Font(bold=True, size=12, color="366092")
    ws_resumo['A11'].font = Font(bold=True, size=12, color="366092")
    
    # Estilizar coluna de labels
    for row in range(2, ws_resumo.max_row + 1):
        ws_resumo[f'A{row}'].font = Font(bold=True)
    
    ajustar_largura_colunas(ws_resumo)
    
    # ===== ABA 2: RESUMO POR PRODUTO =====
    ws_produtos = wb.create_sheet("Resumo por Produto")
    
    # Agrupar por produto
    produtos_agrupados = {}
    for item in itens_validos:
        id_produto = item.get('id_produto')
        quantidade = item.get('quantidade', 0)
        
        if id_produto in produtos_agrupados:
            produtos_agrupados[id_produto]['quantidade'] += quantidade
            produtos_agrupados[id_produto]['ocorrencias'] += 1
        else:
            produtos_agrupados[id_produto] = {
                'quantidade': quantidade,
                'ocorrencias': 1
            }
    
    # Criar cabeçalho
    ws_produtos.append(["ID Produto", "Quantidade Total", "Ocorrências", "Média por Ocorrência"])
    estilizar_cabecalho(ws_produtos)
    
    # Adicionar dados
    for id_produto in sorted(produtos_agrupados.keys()):
        info = produtos_agrupados[id_produto]
        media = round(info['quantidade'] / info['ocorrencias'], 2)
        ws_produtos.append([
            id_produto,
            info['quantidade'],
            info['ocorrencias'],
            media
        ])
    
    ajustar_largura_colunas(ws_produtos)
    
    # ===== ABA 3: DETALHAMENTO DE ITENS =====
    ws_itens = wb.create_sheet("Detalhamento de Itens")
    
    # Criar cabeçalho
    ws_itens.append([
        "Item #",
        "ID Produto",
        "Descrição",
        "Quantidade",
        "Usar Lista Preço",
        "Manter Estrutura",
        "IDs Distribuição",
        "Componentes",
        "Perguntas Gerais"
    ])
    estilizar_cabecalho(ws_itens)
    
    # Adicionar dados
    for idx, item in enumerate(itens_validos, 1):
        ws_itens.append([
            idx,
            item.get('id_produto', 'N/A'),
            item.get('descricao', 'N/A'),
            item.get('quantidade', 0),
            'Sim' if item.get('usar_listapreco') == 1 else 'Não',
            'Sim' if item.get('manter_estrutura_mod_produto') == 1 else 'Não',
            len(item.get('ids_distribuicao', [])),
            len(item.get('componentes', [])),
            len(item.get('perguntas_gerais', []))
        ])
    
    ajustar_largura_colunas(ws_itens)
    
    # ===== ABA 4: ANÁLISE DE DISTRIBUIÇÃO =====
    ws_distribuicao = wb.create_sheet("Análise de Distribuição")
    
    # Criar cabeçalho
    ws_distribuicao.append([
        "Item #",
        "Descrição",
        "Quantidade de IDs",
        "Primeiro ID",
        "Último ID"
    ])
    estilizar_cabecalho(ws_distribuicao)
    
    # Adicionar dados
    for idx, item in enumerate(itens_validos, 1):
        ids_dist = item.get('ids_distribuicao', [])
        ws_distribuicao.append([
            idx,
            item.get('descricao', 'N/A'),
            len(ids_dist),
            ids_dist[0] if ids_dist else 'N/A',
            ids_dist[-1] if ids_dist else 'N/A'
        ])
    
    ajustar_largura_colunas(ws_distribuicao)
    
    # ===== ABA 5: PERGUNTAS POR ITEM =====
    ws_perguntas = wb.create_sheet("Perguntas por Item")
    
    # Criar cabeçalho
    ws_perguntas.append([
        "Item #",
        "Descrição",
        "ID Pergunta",
        "Resposta"
    ])
    estilizar_cabecalho(ws_perguntas)
    
    # Adicionar dados
    for idx, item in enumerate(itens_validos, 1):
        perguntas = item.get('perguntas_gerais', [])
        if perguntas:
            for pergunta in perguntas:
                if pergunta:  # Verificar se não é vazio
                    ws_perguntas.append([
                        idx,
                        item.get('descricao', 'N/A'),
                        pergunta.get('id_pergunta', 'N/A'),
                        pergunta.get('resposta', 'N/A')
                    ])
        else:
            ws_perguntas.append([
                idx,
                item.get('descricao', 'N/A'),
                'Sem perguntas',
                ''
            ])
    
    ajustar_largura_colunas(ws_perguntas)
    
    # Definir nome do arquivo de saída
    if not arquivo_saida:
        arquivo_saida = f"relatorio_pageflow_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Salvar o arquivo
    wb.save(arquivo_saida)
    print(f"✓ Relatório Excel gerado com sucesso em: {arquivo_saida}")
    print(f"\nAbas criadas:")
    print("  - Resumo Geral")
    print("  - Resumo por Produto")
    print("  - Detalhamento de Itens")
    print("  - Análise de Distribuição")
    print("  - Perguntas por Item")
    
    return arquivo_saida


def main():
    """Função principal"""
    import sys
    
    # Verificar argumentos
    if len(sys.argv) < 2:
        print("Uso: python gerar_relatorio_excel.py <arquivo.json> [arquivo_saida.xlsx]")
        print("\nExemplos:")
        print("  python gerar_relatorio_excel.py .json")
        print("  python gerar_relatorio_excel.py .json relatorio.xlsx")
        sys.exit(1)
    
    arquivo_json = sys.argv[1]
    arquivo_saida = sys.argv[2] if len(sys.argv) > 2 else None
    
    # Verificar se o arquivo existe
    if not Path(arquivo_json).exists():
        print(f"✗ Erro: Arquivo '{arquivo_json}' não encontrado!")
        sys.exit(1)
    
    # Gerar relatório
    try:
        gerar_relatorio_excel(arquivo_json, arquivo_saida)
    except Exception as e:
        print(f"✗ Erro ao gerar relatório: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
